import boto3
from datetime import datetime, timedelta, timezone
import subprocess
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Replace placeholder values with your AWS credentials
def get_aws_credentials():
    AWS_ACCESS_KEY = input("Enter your AWS Access Key: ")
    AWS_SECRET_KEY = input("Enter your AWS Secret Key: ")
    AWS_REGION = input("Enter your AWS Region: ")
    return AWS_ACCESS_KEY, AWS_SECRET_KEY, AWS_REGION

# Initialize AWS clients
def initialize_aws_clients(AWS_ACCESS_KEY, AWS_SECRET_KEY, AWS_REGION):
    try:
        ec2_client = boto3.client('ec2', aws_access_key_id=AWS_ACCESS_KEY, aws_secret_access_key=AWS_SECRET_KEY, region_name=AWS_REGION)
        elbv2_client = boto3.client('elbv2', aws_access_key_id=AWS_ACCESS_KEY, aws_secret_access_key=AWS_SECRET_KEY, region_name=AWS_REGION)
        cloudwatch_client = boto3.client('cloudwatch', aws_access_key_id=AWS_ACCESS_KEY, aws_secret_access_key=AWS_SECRET_KEY, region_name=AWS_REGION)
        eks_client = boto3.client('eks', aws_access_key_id=AWS_ACCESS_KEY, aws_secret_access_key=AWS_SECRET_KEY, region_name=AWS_REGION)
        ecs_client = boto3.client('ecs', aws_access_key_id=AWS_ACCESS_KEY, aws_secret_access_key=AWS_SECRET_KEY, region_name=AWS_REGION)
        return ec2_client, elbv2_client, cloudwatch_client, eks_client, ecs_client

    except Exception as e:
        print(f"Error initializing AWS clients: {str(e)}")
        return None, None, None, None, None


def get_instance_id_from_user():
    instance_id = input("Enter the Instance ID: ")
    return instance_id

def describe_running_instances(ec2_client):
    print("----- Describe Running Instances -----")
    try:
        response = ec2_client.describe_instances(Filters=[{'Name': 'instance-state-name', 'Values': ['running']}])
        for reservation in response['Reservations']:
            for instance in reservation['Instances']:
                print(f"Instance ID: {instance['InstanceId']}")
                print(f"Instance Type: {instance['InstanceType']}")
                print(f"AMI ID: {instance['ImageId']}")
                print(f"Tags: {instance.get('Tags', 'No tags')}")
                print("------------------------------")
    except Exception as e:
        print(f"Error describing instances: {str(e)}")
def describe_instances(instance_ids=None):
    print("---find unused resources of EC2---")
    # AWS CLI command to describe instances
    aws_cli_command = "aws ec2 describe-instances"
    # If specific instance IDs are provided, include them in the command
    if instance_ids:
        instance_ids_str = " ".join(instance_ids)
        aws_cli_command += f" --instance-ids {instance_ids_str}"

    # Execute the AWS CLI command using subprocess
    result = subprocess.run(aws_cli_command, shell=True, capture_output=True, text=True)

    # Parse the JSON output and return the relevant data
    if result.returncode == 0:
        output_json = json.loads(result.stdout)
        return output_json.get("Reservations", [])
    else:
        print(f"Failed to describe instances. Error: {result.stderr}")
        return None
def list_unused_security_groups():

    regions = regions()
    for region in regions:
        print(f"\nRunning command for region: {region}")

        # Example: Replace 'your-instance-id' with the actual ID of your EC2 instance
        instance_ids_to_describe = ['your-instance-id']

        # Get information about EC2 instances
        instances_info = describe_instances(instance_ids_to_describe)

        if instances_info:
            print("Instances Information:")
            for reservation in instances_info:
                for instance in reservation.get("Instances", []):
                    instance_id = instance.get("InstanceId")
                    instance_state = instance.get("State", {}).get("Name")
                    print(f"Instance ID: {instance_id}, State: {instance_state}")
        else:
            print("No instance information available.")

        ec2_client = boto3.client('ec2', region_name=region)
        response = ec2_client.describe_security_groups(
            Filters=[{'Name': 'vpc-id', 'Values': ['']}],  # Filter security groups not attached to any VPC
        )

        unused_security_groups = [sg['GroupId'] for sg in response['SecurityGroups']]
        print(f"Unused Security Groups in {region}: {unused_security_groups}")

if __name__ == "__main__":
    list_unused_security_groups()

def check_ami_encryption(ec2_client, ami_id):
    print("----- Check AMI Encryption -----")
    try:
        response = ec2_client.describe_images(ImageIds=[ami_id])
        ami = response['Images'][0]
        encryption_status = ami.get('BlockDeviceMappings', [{}])[0].get('Ebs', {}).get('Encrypted', False)
        print(f"Encryption Status for AMI {ami_id}: {'Encrypted' if encryption_status else 'Not Encrypted'}")
    except Exception as e:
        print(f"Error checking AMI encryption: {str(e)}")

def describe_images_command():
    print("----- Describe Images Command -----")
    # Modify the command based on your requirements
    try:
        describe_images_output = subprocess.check_output(['aws', 'ec2', 'describe-images', '--output', 'json'])
        images = json.loads(describe_images_output)
        latest_image = max(images['Images'], key=lambda x: x['CreationDate'])
        print(f"Latest Image ID: {latest_image['ImageId']}")
        print(f"Image Creation Date: {latest_image['CreationDate']}")
        print(f"Image Tags: {latest_image.get('Tags', 'No tags')}")
    except Exception as e:
        print(f"Error checking images command: {str(e)}")

def fetch_cpu_utilization(instance_id):
    print("----- Fetch CPU Utilization -----")
    end_time = datetime.utcnow()
    start_time = end_time - timedelta(days=30)
    response = cloudwatch_client.get_metric_data(
        MetricDataQueries=[
            {
                'Id': 'm1',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/EC2',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'InstanceId',
                                'Value': instance_id
                            },
                        ]
                    },
                    'Period': 3600,
                    'Stat': 'Maximum',
                },
                'ReturnData': True,
            },
        ],
        StartTime=start_time,
        EndTime=end_time,
    )
    metric_data = response['MetricDataResults'][0]['Values']
    print(f"CPU Utilization for Instance {instance_id} in the last 30 days: {metric_data}")

def peak_cpu_usage(instance_id):
    print("----- Peak CPU Usage -----")
    end_time = datetime.utcnow()
    start_time = end_time - timedelta(days=30)
    response = cloudwatch_client.get_metric_data(
        MetricDataQueries=[
            {
                'Id': 'm1',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/EC2',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'InstanceId',
                                'Value': instance_id
                            },
                        ]
                    },
                    'Period': 3600,
                    'Stat': 'Maximum',
                },
                'ReturnData': True,
            },
        ],
        StartTime=start_time,
        EndTime=end_time,
    )
    metric_data = response['MetricDataResults'][0]['Values']
    peak_cpu = max(metric_data)
    print(f"Peak CPU Usage for Instance {instance_id} in the last 30 days: {peak_cpu}%")

def check_elastic_load_balancer_attachment(instance_id):
    print("----- Check Elastic Load Balancer Attachment -----")
    response = elbv2_client.describe_target_health(TargetGroupArn='your-target-group-arn', Targets=[{'Id': instance_id}])
    attached = len(response['TargetHealthDescriptions']) > 0
    print(f"Attached to Elastic Load Balancer: {'Yes' if attached else 'No'}")

def elastic_ip_configuration_check(instance_id):
    print("----- Elastic IP Configuration Check -----")
    response = ec2_client.describe_addresses(Filters=[{'Name': 'instance-id', 'Values': [instance_id]}])
    elastic_ip_attached = len(response['Addresses']) > 0
    print(f"Elastic IP Attached: {'Yes' if elastic_ip_attached else 'No'}")

def subnet_type_check(instance_id):
    print("----- Subnet Type Check -----")
    response = ec2_client.describe_instances(InstanceIds=[instance_id])
    subnet_id = response['Reservations'][0]['Instances'][0]['SubnetId']
    response_subnet = ec2_client.describe_subnets(SubnetIds=[subnet_id])
    subnet_type = response_subnet['Subnets'][0]['MapPublicIpOnLaunch']
    print(f"Subnet Type for Instance {instance_id}: {'Public' if subnet_type else 'Private'}")

def elastic_ip_configuration(instance_id):
    print("----- Elastic IP Configuration Check -----")
    response = ec2_client.describe_addresses(Filters=[{'Name': 'instance-id', 'Values': [instance_id]}])
    elastic_ip_attached = len(response['Addresses']) > 0
    print(f"Elastic IP Attached: {'Yes' if elastic_ip_attached else 'No'}")

def subnet_type_check(instance_id):
    print("----- Subnet Type Check -----")
    response = ec2_client.describe_instances(InstanceIds=[instance_id])
    subnet_id = response['Reservations'][0]['Instances'][0]['SubnetId']
    response_subnet = ec2_client.describe_subnets(SubnetIds=[subnet_id])
    subnet_type = response_subnet['Subnets'][0]['MapPublicIpOnLaunch']
    print(f"Subnet Type for Instance {instance_id}: {'Public' if subnet_type else 'Private'}")

def check_network_acl_and_security_groups(instance_id):
    print("----- Check Network ACLs and Security Groups -----")
    response = ec2_client.describe_instances(InstanceIds=[instance_id])
    security_groups = response['Reservations'][0]['Instances'][0]['SecurityGroups']
    network_acl_id = response['Reservations'][0]['Instances'][0]['NetworkInterfaces'][0]['NetworkAclId']
    print(f"Security Groups for Instance {instance_id}: {security_groups}")
    print(f"Network ACL ID for Instance {instance_id}: {network_acl_id}")

def monitoring_optimization(instance_id):
    print("----- Monitoring Optimization Check -----")
    response = ec2_client.describe_instances(InstanceIds=[instance_id])
    detailed_monitoring_enabled = response['Reservations'][0]['Instances'][0].get('Monitoring', {}).get('State', 'disabled')
    print(f"Detailed Monitoring Enabled for Instance {instance_id}: {'Yes' if detailed_monitoring_enabled == 'enabled' else 'No'}")

def placement_group_check(instance_id):
    print("----- Placement Group Check -----")
    response = ec2_client.describe_instances(InstanceIds=[instance_id])
    placement_group_name = response['Reservations'][0]['Instances'][0].get('Placement', {}).get('GroupName', 'Not in a Placement Group')
    print(f"Placement Group for Instance {instance_id}: {placement_group_name}")

def storage_performance_check(instance_id):
    print("----- Storage Performance Check -----")
    response = ec2_client.describe_instances(InstanceIds=[instance_id])
    block_device_mappings = response['Reservations'][0]['Instances'][0].get('BlockDeviceMappings', [])
    for block_device in block_device_mappings:
        volume_id = block_device['Ebs']['VolumeId']
        response_volume = ec2_client.describe_volumes(VolumeIds=[volume_id])
        volume_type = response_volume['Volumes'][0]['VolumeType']
        print(f"Storage Type for Volume {volume_id}: {volume_type}")

def auto_scaling_group_check(instance_id):
    print("----- Auto Scaling Group Check -----")
    response = ec2_client.describe_auto_scaling_instances(InstanceIds=[instance_id])
    auto_scaling_group_names = [instance['AutoScalingGroupName'] for instance in response['AutoScalingInstances']]
    print(f"Auto Scaling Groups for Instance {instance_id}: {auto_scaling_group_names}")

def availability_zone_and_region_check(instance_id):
    print("----- Availability Zone and Region Check -----")
    response = ec2_client.describe_instances(InstanceIds=[instance_id])
    availability_zone = response['Reservations'][0]['Instances'][0]['Placement']['AvailabilityZone']
    region = availability_zone[:-1]
    print(f"Availability Zone for Instance {instance_id}: {availability_zone}")
    print(f"Region for Instance {instance_id}: {region}")

def cost_monitoring(instance_id):
    print("----- Cost Monitoring -----")
    response = ec2_client.describe_instances(InstanceIds=[instance_id])
    instance_type = response['Reservations'][0]['Instances'][0]['InstanceType']

    # Calculate estimated monthly cost
    response_pricing = ec2_client.describe_spot_price_history(
        InstanceTypes=[instance_type],
        MaxResults=1,
        ProductDescriptions=['Linux/UNIX'],
    )
    price_per_hour = float(response_pricing['SpotPriceHistory'][0]['SpotPrice'])
    hours_in_month = 24 * 30  # Assuming a month with 30 days
    estimated_monthly_cost = price_per_hour * hours_in_month
    
    append_to_excel_data("Cost Monitoring - Estimated Cost per Hour", price_per_hour)
    append_to_excel_data("Cost Monitoring - Estimated Monthly Cost", estimated_monthly_cost)
    append_to_excel_data("Cost Monitoring - Compute Cost", compute_cost)
    append_to_excel_data("Cost Monitoring - Storage Cost", storage_cost)


    print(f"Estimated Cost per Hour for Instance {instance_id} ({instance_type}): {price_per_hour}")
    print(f"Estimated Monthly Cost: {estimated_monthly_cost}")
    

    # Cost breakdown (example: separate compute and storage costs)
    compute_cost = price_per_hour * hours_in_month  # Modify based on your specific pricing model
    storage_cost = 0  # Add logic to calculate storage cost if applicable
    print(f"Compute Cost: {compute_cost}")
    print(f"Storage Cost: {storage_cost}")

def describe_images_command():
    print("----- Describe Images Command -----")
    # Modify the command based on your requirements
    describe_images_output = subprocess.check_output(['aws', 'ec2', 'describe-images', '--output', 'json'])
    images = json.loads(describe_images_output)
    print(f"Images: {images}")

def list_custom_amis():
    print("----- List Custom AMIs -----")
    response = ec2_client.describe_images(Filters=[{'Name': 'owner-id', 'Values': ['self']}])
    custom_amis = response['Images']
    print(f"Custom AMIs: {custom_amis}")

def ami_identification(instance_id):
    print("----- AMI Identification -----")
    response = ec2_client.describe_instances(InstanceIds=[instance_id])
    ami_id = response['Reservations'][0]['Instances'][0]['ImageId']
    print(f"AMI ID for Instance {instance_id}: {ami_id}")

#improved code by usinf if else(can be made changes in same way if needed)
def get_instances_by_state(ec2_client, state, placeholder="Enter the Instance ID"):
    print(f"----- Get Instances by State ({state.capitalize()}) -----")
    instance_id = input(f"{placeholder}: ")
    response = ec2_client.describe_instances(Filters=[{'Name': 'instance-state-name', 'Values': [state]}])

    instances = response['Reservations'][0]['Instances'] if response['Reservations'] else []

    if instances:
        for instance in instances:
            print(f"Instance ID: {instance['InstanceId']}")
            print(f"Instance Type: {instance['InstanceType']}")
            print(f"AMI ID: {instance['ImageId']}")
            print(f"Tags: {instance.get('Tags', 'No tags')}")
            print("------------------------------")
    else:
        print(f"No instances found in {state} state.")

def attached_to_elastic_load_balancer(instance_id):
    print("----- Attached to Elastic Load Balancer -----")
    response = elbv2_client.describe_target_health(TargetGroupArn='your-target-group-arn', Targets=[{'Id': instance_id}])
    attached = len(response['TargetHealthDescriptions']) > 0
    print(f"Attached to Elastic Load Balancer: {'Yes' if attached else 'No'}")

def elastic_ip_configuration_check(instance_id):
    print("----- Elastic IP Configuration Check -----")
    response = ec2_client.describe_addresses(Filters=[{'Name': 'instance-id', 'Values': [instance_id]}])
    elastic_ip_attached = len(response['Addresses']) > 0
    print(f"Elastic IP Attached: {'Yes' if elastic_ip_attached else 'No'}")

def elastic_ip_association_check(instance_id):
    print("----- Elastic IP Association Check -----")
    response = ec2_client.describe_addresses(Filters=[{'Name': 'instance-id', 'Values': [instance_id]}])
    elastic_ip_associated = len(response['Addresses']) > 0
    print(f"Elastic IP Associated: {'Yes' if elastic_ip_associated else 'No'}")

def public_dns_check(instance_id):
    print("----- Public DNS Check -----")
    response = ec2_client.describe_instances(InstanceIds=[instance_id])
    public_dns = response['Reservations'][0]['Instances'][0].get('PublicDnsName', 'Not available')
    print(f"Public DNS for Instance {instance_id}: {public_dns}")

def describe_running_eks_clusters():
    print("----- Describe Running EKS Clusters -----")
    response = eks_client.list_clusters()
    for cluster_name in response['clusters']:
        print(f"EKS Cluster Name: {cluster_name}")

        # Additional details
        cluster_details = eks_client.describe_cluster(name=cluster_name)
        print(f"EKS Cluster Version: {cluster_details['cluster']['version']}")
        print(f"EKS Cluster Node IAM Role: {cluster_details['cluster']['roleArn']}")
        

def eks_nodes_health_check(cluster_name):
    print("----- EKS Nodes Health Check -----")
    response = eks_client.describe_nodegroup(
        clusterName=cluster_name,
        nodegroupName='your-node-group-name'  # Replace with your node group name
    )
    nodegroup_status = response['nodegroup']['status']
    print(f"Node Group Status: {nodegroup_status}")

    # Check for node health, adjust as needed based on your criteria
    if nodegroup_status == 'ACTIVE':
        print("Node Group is active. Checking node health...")
        nodes = eks_client.list_nodes(clusterName=cluster_name)
        for node in nodes['nodes']:
            node_health = node['health']['status']
            print(f"Node ID: {node['id']}, Health Status: {node_health}")
    else:
        print("Node Group is not active.")

def describe_running_ecs_clusters():
    print("----- Describe Running ECS Clusters -----")
    response = ecs_client.list_clusters()
    for cluster_arn in response['clusterArns']:
        print(f"ECS Cluster ARN: {cluster_arn}")

        #Cluster Instances
        print("----- Cluster Instances -----")
        instances = ecs_client.list_container_instances(cluster=cluster_arn)
        for instance_arn in instances['containerInstanceArns']:
            instance_details = ecs_client.describe_container_instances(cluster=cluster_arn, containerInstances=[instance_arn])
            instance = instance_details['containerInstances'][0]
            ecs_instance_data.append(f"Instance ID: {instance['ec2InstanceId']}")
            ecs_instance_data.append(f"Status: {instance['status']}")
            ecs_instance_data.append("------------------------------")
            ecs_cluster_data.append("----- Cluster Instances -----")
            ecs_cluster_data.extend(ecs_instance_data)

        # Cluster Services
        ecs_service_data = []  # Data specific to ECS services
        services = ecs_client.list_services(cluster=cluster_arn)
        for service_arn in services['serviceArns']:
            service_details = ecs_client.describe_services(cluster=cluster_arn, services=[service_arn])
            service = service_details['services'][0]
            ecs_service_data.append(f"Service Name: {service['serviceName']}")
            ecs_service_data.append(f"Desired Tasks: {service['desiredCount']}")
            ecs_service_data.append(f"Running Tasks: {service['runningCount']}")
            ecs_service_data.append("------------------------------")
        
        ecs_cluster_data.append("----- Cluster Services -----")
        ecs_cluster_data.extend(ecs_service_data)
        
        # Append ECS cluster data to excel_report_data
        append_to_excel_data("Describe Running ECS Clusters", ecs_cluster_data)
def describe_ebs_volumes():
    print("----- Describe EBS Volumes -----")
    response = ec2_client.describe_volumes()
    for volume in response['Volumes']:
        print(f"EBS Volume ID: {volume['VolumeId']}")
        print(f"Size: {volume['Size']} GB")
        print(f"Status: {volume['State']}")
        print("------------------------------")
# Combine the two describe_ebs_volumes() functions into one
# Define AWS EC2 client

ec2_client = boto3.client('ec2')
def describe_unused_ebs_volumes():
    print("----- Describe Unused_EBS Volumes -----")
    response = ec2_client.describe_volumes()
    
    for volume in response['Volumes']:
        print(f"EBS Volume ID: {volume['VolumeId']}")
        print(f"Size: {volume['Size']} GB")
        print(f"Status: {volume['State']}")

        # Volume Attachments
        print("----- Volume Attachments -----")
        for attachment in volume['Attachments']:
            print(f"Instance ID: {attachment['InstanceId']}")
            print(f"Device: {attachment['Device']}")
            print(f"Attachment State: {attachment['State']}")
            print("------------------------------")

        # Volume Encryption
        print("----- Volume Encryption -----")
        encryption_status = volume.get('Encrypted', False)
        print(f"Encryption Status: {'Encrypted' if encryption_status else 'Not Encrypted'}")
        print("------------------------------")

def menu():
    print("Choose an action:")
    print("1. List Unused Security Groups")
    print("2. Check EBS Snapshots Older Than 30 Days")
    print("3. Describe EBS Volumes")
    print("4. Quit")

    choice = input("Enter your choice: ")

    if choice == '1':
        list_unused_security_groups()
    elif choice == '2':
        check_ebs_snapshots_older_than_30_days()
    elif choice == '3':
        describe_ebs_volumes()
    elif choice == '4':
        print("Exiting the script.")
        exit()
    else:
        print("Invalid choice. Please enter a valid option.")

def check_ebs_snapshots_older_than_30_days():
    print("----- Check EBS Snapshots Older Than 30 Days -----")
    snapshot_response = ec2_client.describe_snapshots(OwnerIds=['self'])
    current_time = datetime.now(timezone.utc)

    for snapshot in snapshot_response['Snapshots']:
        snapshot_id = snapshot['SnapshotId']
        volume_id = snapshot['VolumeId']
        volume_size = snapshot['VolumeSize']
        start_time = snapshot['StartTime']
        days_old = (current_time - start_time).days

        print(f"Snapshot ID: {snapshot_id}")
        print(f"Volume ID: {volume_id}")
        print(f"Snapshot Age: {days_old} days")

        if days_old > 30:
            # Fetch additional information about the volume
            volume_response = ec2_client.describe_volumes(VolumeIds=[volume_id])
            volume = volume_response['Volumes'][0]

            instance_id = ""
            volume_name = ""
            volume_size = volume['Size']
            volume_type = volume['VolumeType']

            # Check if the volume has attachments (i.e., associated with an instance)
            if volume['Attachments']:
                instance_id = volume['Attachments'][0]['InstanceId']

                # Fetch additional information about the instance
                instance_response = ec2_client.describe_instances(InstanceIds=[instance_id])
                instance = instance_response['Reservations'][0]['Instances'][0]
                volume_name = instance.get('Tags', [{'Key': 'Name', 'Value': 'N/A'}])[0]['Value']

            print(f"Instance ID: {instance_id}")
            print(f"Volume Name: {volume_name}")
            print(f"Volume Size: {volume_size} GB")
            print(f"Volume Type: {volume_type}")
            print("Consider reviewing and deleting if necessary.")
            print("------------------------------")

            append_to_excel_data("EBS Snapshots - Snapshot ID", snapshot_id)
            append_to_excel_data("EBS Snapshots - Volume ID", volume_id)
            append_to_excel_data("EBS Snapshots - Snapshot Age", f"{days_old} days")
            append_to_excel_data("EBS Snapshots - Instance ID", instance_id)
            append_to_excel_data("EBS Snapshots - Volume Name", volume_name)
            append_to_excel_data("EBS Snapshots - Volume Size", f"{volume_size} GB")
            append_to_excel_data("EBS Snapshots - Volume Type", volume_type)

        
def describe_running_ecs_clusters(cluster_arn):
    print("----- Describe Running ECS Clusters -----")
    response = ecs_client.list_clusters()
    for cluster_arn in response['clusterArns']:
        print(f"ECS Cluster ARN: {cluster_arn}")

        # Cluster Instances
        print("----- Cluster Instances -----")
        instances = ecs_client.list_container_instances(cluster=cluster_arn)
        for instance_arn in instances['containerInstanceArns']:
            instance_details = ecs_client.describe_container_instances(cluster=cluster_arn, containerInstances=[instance_arn])
            instance = instance_details['containerInstances'][0]
            print(f"Instance ID: {instance['ec2InstanceId']}")
            print(f"Status: {instance['status']}")
            print("------------------------------")

        # Cluster Services
        print("----- Cluster Services -----")
        services = ecs_client.list_services(cluster=cluster_arn)
        for service_arn in services['serviceArns']:
            service_details = ecs_client.describe_services(cluster=cluster_arn, services=[service_arn])
            service = service_details['services'][0]
            print(f"Service Name: {service['serviceName']}")
            print(f"Desired Tasks: {service['desiredCount']}")
            print(f"Running Tasks: {service['runningCount']}")
            print("------------------------------")

def create_excel_report(data, filename='report.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write headers
    headers = ["Instance ID", "Instance Type", "AMI ID", "Tags"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = Font(bold=True)

    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}{row_num}"] = cell_value

    # Save the Excel workbook
    wb.save(filename)
    print(f"Excel report saved to {filename}")

 def append_to_excel_data(header, data):
    # Helper function to append data to the excel_report_data list
    excel_report_data.append([header, data])   

# replace 'your-instance-id' and 'your-ami-id' with actual values,
# Prompt for the instance ID
if __name__ == "__main__":
    # Initialize AWS clients
    AWS_ACCESS_KEY, AWS_SECRET_KEY, AWS_REGION = get_aws_credentials()
    ec2_client, elbv2_client, cloudwatch_client, eks_client, ecs_client = initialize_aws_clients(AWS_ACCESS_KEY, AWS_SECRET_KEY, AWS_REGION)

    # Prompt for the instance ID
    instance_id = input("Enter the Instance ID: ")

    # Function calls
    describe_running_instances(ec2_client)
    check_ami_encryption(ec2_client, instance_id)
    # ... (other function calls)

    describe_images_command()
    list_custom_amis()
    ami_identification(instance_id)
    get_instances_by_state('running', "Enter the Instance ID")
    attached_to_elastic_load_balancer(instance_id)
    elastic_ip_configuration_check(instance_id)
    

    # Additional function calls
    elastic_ip_association_check(instance_id)
    public_dns_check(instance_id)
    eks_nodes_health_check('your-cluster-name', "Enter the cluster ID" )  # Replace 'your-cluster-name' with the actual EKS cluster name
    describe_running_eks_clusters()
    describe_running_ecs_clusters()
    describe_ebs_volumes()
    
    # Cost monitoring and other checks
    cost_monitoring(instance_id)
    check_ebs_snapshots_older_than_30_days()  # Assuming the function is defined

    # Collect data for the Excel report
    excel_report_data = []

    # Example data for one instance, replace this with actual data from your functions
    instance_data = [
        "Instance_ID_Value",
        "Instance_Type_Value",
        "AMI_ID_Value",
        "Tags_Value",
    ]
    excel_report_data.append(instance_data)
    # Example usage of append_to_excel_data
append_to_excel_data("Header for Cost Monitoring", "Cost Monitoring data goes here")
append_to_excel_data("Header for EBS Snapshots", "EBS Snapshots data goes here")
append_to_excel_data("Describe Running EKS Clusters", eks_cluster_data)
# Example usage of append_to_excel_data
append_to_excel_data("Header for EKS Clusters", "EKS Clusters data goes here")
append_to_excel_data("Header for ECS Clusters", "ECS Clusters data goes here")

    # Call the create_excel_report function
    create_excel_report(excel_report_data)
